import ldap3
from dotenv import load_dotenv
import pandas as pd
import os
import openpyxl

def get_all_users(username, password, server_url, base_dn, empID):
    # Establish connection to LDAP server
    server = ldap3.Server(server_url)
    connection = ldap3.Connection(server, user=username, password=password, auto_bind=True)

    # Search for users with the specified last name
    search_filter = '(&(objectClass=user)(sn={}))'.format(empID)
    connection.search(search_base=base_dn,
                      search_filter=search_filter,
                      attributes=ldap3.ALL_ATTRIBUTES)

    # Retrieve users and their attributes
    users_with_empID = []
    for entry in connection.entries:
        
        users_with_empID.append(entry)

    # Close connection
    connection.unbind()

    return users_with_empID

# Locate the column that's the empID to search the LDAP
def find_column_index(sheet, column_label):
    for col_idx in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col_idx).value
        if cell_value == column_label:
            return col_idx
    return None


def read_PSA_report(filename, empID_label):
    # Open the workbook
    wb = openpyxl.load_workbook(filename)
    
    # Select the first sheet
    sheet = wb.active
    
    # Find the column index for 'empID' label
    empID_column = find_column_index(sheet, empID_label)
    if empID_column is None:
        print(f"Column '{empID_label}' not found.")
        return
    
    # Iterate through rows in the sheet
    for row in sheet.iter_rows(values_only=True):
        # Print empID column
        empID = row[empID_column - 1]
        print("Employee ID:", empID)

    # Close the workbook
    wb.close()

def combine_Records():
    return


# LDAP server configuration
load_dotenv()
ldap_username = os.environ.get('ldap_username')
ldap_password = os.environ.get('ldap_password')
ldap_server_url = os.environ.get('ldap_server_url')
ldap_base_dn = os.environ.get('ldap_base_dn')
psa_report_path = os.environ.get('psa_report_path')
output_file_path = os.environ.get('output_path')
ldap_search_filter = '(employeeID={})'

#read_PSA_report(psa_report_path, 'EmplID')

psa_report = pd.read_excel(psa_report_path, engine='openpyxl')

#users = get_all_users(ldap_username, ldap_password, ldap_server_url, ldap_base_dn, empID)


# Iterate over the excel file
for index, row in psa_report:
    employee_id = row['EmplID']
    search_filter= ldap_search_filter.format(employee_id)
    print(search_filter)


